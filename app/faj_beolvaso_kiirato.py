import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook
from pathlib import Path
import requests

class BeolvasKiirat:
    def __init__(self):
        self.mappa = ""

    def csv_beolvasasa_databol(self, nev):
        input_file = "data/" + nev
        df = pd.read_csv(input_file, encoding="utf-8", sep=";")
        return df

    def csv_beolvas_df(self):
        root = tk.Tk()
        root.withdraw()  # főablak elrejtése
        file_path = filedialog.askopenfilename(
            title="Válassz CSV fájlt",
            filetypes=[("CSV fájlok", "*.csv"), ("Minden fájl", "*.*")]
        )
        if not file_path:
            print("Nem választottál fájlt.")
            return None

        self.mappa = file_path
        df = pd.read_csv(file_path, sep=";")
        print(f"Beolvasva: {file_path} ({len(df)} sor)")
        return df

    def szerkesztett_excel_beolvaso(self, fajlnev, data_mappa="data"):
        """
        Beolvassa a sablon Excel fájlt a data mappából, megtartva a formázást.
        - fajlnev: a fájl neve pl. "arajanlat.xlsx"
        - data_mappa: a mappa ahol a fájl van (alapértelmezett: 'data')
        """
        # Teljes útvonal összeállítása
        fajl_ut = os.path.join(data_mappa, fajlnev)

        if not os.path.exists(fajl_ut):
            print(f"Hiba: a fájl nem található: {fajl_ut}")
            return None, None

        print(f"Beolvasás: {fajl_ut}")

        # openpyxl-lel beolvassuk a sablont úgy, hogy a stílusok megmaradjanak
        wb = load_workbook(fajl_ut, data_only=False)  # data_only=False -> stílusok megmaradnak

        # első lap kiválasztása
        ws = wb[wb.sheetnames[0]]  # vagy konkrét név: wb['Árajánlat']

        return wb, ws

    def df_kiiratasa_exelbe(self, df, nev, mappa):
        # Ha a mappa nem létezik, létrehozzuk
        os.makedirs(mappa, exist_ok=True)
        # Teljes fájlútvonal összeállítása
        fajl_ut = os.path.join(mappa, nev)
        # Mentés Excelbe
        df.to_excel(fajl_ut, index=False)

        print(f"✅ Fájl sikeresen mentve ide: {fajl_ut}")

    def exel_kiiratasa(self, wb, utvonal, teljes_fajlut=True, nev="arajanlat.xlsx"):
        """
        Excel fájl mentése.
        - wb: openpyxl Workbook
        - utvonal:
            - ha teljes_fajlut=True → teljes fájlút
            - ha teljes_fajlut=False → célmappa
        - nev: fájlnév (csak akkor számít, ha nem teljes a fájlút)
        """

        if teljes_fajlut:
            fajl_ut = utvonal
            mappa = os.path.dirname(fajl_ut)
        else:
            mappa = utvonal
            fajl_ut = os.path.join(mappa, nev)

        os.makedirs(mappa, exist_ok=True)

        wb.save(fajl_ut)
        print(f"✅ Fájl sikeresen mentve ide: {fajl_ut}")

    def excel_beolvas_df(self, ut: str | Path, munkalap: str | int = 0) -> pd.DataFrame:
        """
        Excel beolvasása DataFrame-be.

        Paraméterek:
          - ut: .xlsx/.xlsm fájl elérési útja (pl. r"C:\\Users\\csiki\\OneDrive\\valami.xlsx")
          - munkalap: munkalap neve (str) vagy indexe (int). Alap: 0 (első lap)
        """
        ut = Path(ut)

        if not ut.exists():
            raise FileNotFoundError(f"Nem találom az Excel fájlt: {ut.resolve()}")

        df = pd.read_excel(ut, sheet_name=munkalap, engine="openpyxl")
        return df

    def excel_beolvas_onedrive_linkbol(
            self,
            megosztasi_link: str,
            cel_mappa: str | Path = "data",
            fajlnev: str = "onedrive_excel.xlsx",
            sheet_name=0,
            **kwargs
    ) -> pd.DataFrame:
        """
        OneDrive Excel letöltése fájlba, majd beolvasás DataFrame-be.
        (Ez kikerüli az xlrd hibát Python 3.12 alatt)
        """

        cel_mappa = Path(cel_mappa)
        cel_mappa.mkdir(parents=True, exist_ok=True)
        fajl_ut = cel_mappa / fajlnev

        # Direkt letöltési link
        if "?download=" not in megosztasi_link:
            if "?" in megosztasi_link:
                download_url = megosztasi_link.split("?")[0] + "?download=1"
            else:
                download_url = megosztasi_link + "?download=1"
        else:
            download_url = megosztasi_link

        # Letöltés fájlba
        r = requests.get(download_url)
        r.raise_for_status()

        with open(fajl_ut, "wb") as f:
            f.write(r.content)

        df = pd.read_excel(
            fajl_ut,
            engine="openpyxl",
            sheet_name=sheet_name,
            **kwargs
        )

        return df

    def utvonalellenorzo(self) -> bool:
        """
        Ellenőrzi a sémát: .../ugyfelek/<ugyfel>/<projekt>/<file>.csv
        Csak a mappastruktúrát nézi, nem a teljes abszolút útvonalat.
        """
        p = Path(self.mappa)

        if p.suffix.lower() != ".csv":
            return False

        parts = [x.lower() for x in p.parts]
        # kell legalább: .../ugyfelek/<ugyfel>/<projekt>/<file>.csv
        if len(parts) < 4:
            return False

        # legyen benne az "ugyfelek" mappa, és utána legyen még 3 elem
        try:
            idx = parts.index("ugyfelek")
        except ValueError:
            return False

        return (len(parts) - idx) >= 4  # ugyfelek, ugyfel, projekt, file.csv

    def faj_tallozo(self, fajnev: str) -> str | None:
        """
        Mentés helyének kiválasztása (Save As).
        Visszaadja a kiválasztott útvonalat, vagy None-t ha a user Cancel.
        """
        root = tk.Tk()
        root.withdraw()  # ne jelenjen meg üres ablak
        root.attributes("-topmost", True)

        path = filedialog.asksaveasfilename(
            title="Hova mentsem az árajánlatot?",
            defaultextension=".xlsx",
            filetypes=[("Excel fájl", "*.xlsx")],
            initialfile=default_filename,
        )

        root.destroy()
        return path if path else None

    def csv_utbol_arajanlat_ut(self) -> str | None:
        """
        Ha a bemeneti útvonal követi a sémát:
          .../ugyfelek/<ugyfel>/<projekt>/<valami>.csv
        akkor:
          .../ugyfelek/<ugyfel>/<projekt>/arajanlat_<projekt>.xlsx

        Ha nem követi:
          felugró tallózó ablakban kiválasztható mentési hely.
        """
        p = Path(self.mappa)

        if self.utvonalellenorzo():
            projektnev = p.parent.name
            out_path = p.parent / f"arajanlat_{projektnev}.xlsx"
            return str(out_path)

        chosen = self.faj_tallozo(fajnev = "arajanlat.xlsx")
        return chosen

    def csv_utbol_arajanlat_ut(self) -> str:
        """
        Bemenet:  .../ugyfelek/<ugyfel>/<projekt>/<valami>.csv
        Kimenet:  .../ugyfelek/<ugyfel>/<projekt>/arajanlat_<projekt>.xlsx
        """
        p = Path(self.mappa)
        projektnev = p.parent.name  # <projekt>
        projekt_mappa = p.parent  # .../<ugyfel>/<projekt>

        out_path = projekt_mappa / f"arajanlat_{projektnev}.xlsx"
        return str(out_path)

