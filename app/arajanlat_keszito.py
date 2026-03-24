import app.szabasjegyzek_szerkeszto as szabjegy
import app.adatgyujto as adatgyujto
import app.faj_beolvaso_kiirato as faj
from openpyxl.drawing.image import Image
import openpyxl.styles as stilus
import os
import pandas as pd
import math

class Arajanlat:
    def __init__(self, ugyfel = "", munkadij_lepesek = "", munkad_erintettanyag=""):
        self.faj = faj.BeolvasKiirat()
        self.df = None
        self.szabjegyszerk = None
        self.ugyfel = ugyfel
        self.munkadijlepesek = munkadij_lepesek
        self.wb, self.ws = self.faj.szerkesztett_excel_beolvaso("minta_arajanlat.xlsx")
        self.sor = 17
        self.kezdosor = self.sor
        self.munkad_erintettanyag = munkad_erintettanyag

    def adatbeolvaso(self):
        self.df = self.faj.csv_beolvas_df()
        self.szabjegyszerk = szabjegy.SzabasjegyzekSzerkeszto(self.df)

    def tablazat_tolto(self, anyagjegyzek):

        for _, row in anyagjegyzek.iterrows():
            self.ws.insert_rows(self.sor)
            self.ws.row_dimensions[self.sor].height = 16

            center_align = stilus.Alignment(vertical="center")

            self.ws.cell(row=self.sor, column=1, value=str(self.sor - self.kezdosor + 1) + ".")
            self.ws.cell(row=self.sor, column=2, value=row.get("Anyag", ""))
            self.ws.cell(row=self.sor, column=4, value=row.get("Szin", ""))

            url = row.get("URL", "")
            if isinstance(url, float) or not url:
                url = ""

            if url:
                c5 = self.ws.cell(row=self.sor, column=5, value="Link")
                c5.hyperlink = str(url)
                c5.style = "Hyperlink"

            menny = row.get("Mennyiseg", "")
            egysegar = row.get("Egysegar", "")

            self.ws.cell(
                row=self.sor,
                column=6,
                value=f"=ROUNDUP({menny},2)"
            )

            self.ws.cell(row=self.sor, column=7, value=row.get("Mertekegyseg", ""))

            c8 = self.ws.cell(row=self.sor, column=8, value=egysegar)
            c8.number_format = '#,##0 "Ft"'

            c9 = self.ws.cell(row=self.sor, column=9, value=f"=F{self.sor}*H{self.sor}")
            c9.number_format = '#,##0 "Ft"'

            self.ws.merge_cells(
                start_row=self.sor, start_column=2, end_row=self.sor, end_column=3
            )

            for col in [1, 2, 4, 6, 7, 8, 9]:
                self.ws.cell(row=self.sor, column=col).alignment = center_align

            self.sor += 1

    def kep_illeszto(self, kep_nev, cella, meret = None):
        kep_ut = os.path.join("data", kep_nev)
        kep = Image(kep_ut)

        if not meret is None:
            kep.width = meret[0]
            kep.height = meret[1]

        self.ws.add_image(kep, cella)

    def telefonszam_formazo(self, v):
        if v is None:
            return ""

        # nan float
        if isinstance(v, float):
            if math.isnan(v):
                return ""
            v = int(v)

        s = str(v).strip()

        if not s:
            return ""

        # ".0" a végén stringként
        if s.endswith(".0"):
            s = s[:-2]

        if s.startswith("+"):
            return s

        return f"+{s}"

    def szemelyes_adatok_onedrive(self):
        if self.ugyfel == "":
            print("Nincs név megadva, a személyes adatok nem lettek beírva!")
            return
        """
        link = "https://1drv.ms/x/c/595ECD328626FCDE/IQCAAzRsxtjJRJhSfNq79RZ8AWpcUCbT_AV30IeTavTymUc?e=lrUfMC"
        ugyfel = (self.faj.excel_beolvas_onedrive_linkbol(megosztasi_link=link, cel_mappa="data", fajlnev="ugyfelek_adat.xlsx"))
        nev = self.vezetek_nev + " " + self.kereszt_nev
        adat = ugyfel[(ugyfel["Név"] == nev)]
        """

        def clean(v):
            return "" if v is None or (isinstance(v, float) and math.isnan(v)) else str(v)

        nev = clean(self.ugyfel.get("Név"))
        email = clean(self.ugyfel.get("Email"))
        tel = clean(self.ugyfel.get("Tel"))
        city = clean(self.ugyfel.get("Város"))
        addr1 = clean(self.ugyfel.get("Lakcím"))
        addr2 = clean(self.ugyfel.get("Emelet, ajtó"))

        cim = f"{city}, {addr1} {addr2}".strip(", ").strip()  # csak akkor rakjon vesszőt, ha van város

        center_align = stilus.Alignment(horizontal="left", vertical="center", wrap_text=True)

        self.ws["G9"] = nev
        self.ws["G9"].alignment = center_align

        self.ws["G10"] = self.telefonszam_formazo(tel)
        self.ws["G10"].alignment = center_align

        self.ws["G11"] = email
        self.ws["G11"].alignment = center_align

        self.ws["G12"] = cim
        self.ws["G12"].alignment = stilus.Alignment(horizontal="left", vertical="top", wrap_text=True)

    def utdij_beirasa(self, toblettszorzo = 1.2):
            # ha F7 üres vagy None, akkor kilépünk
        boltoktav = {
            "Butorkellek": 0,
            "Karnis": 7.3,
            "Borovi": 10.3,
        }
        if self.ugyfel == "":
            print("Nincs cím megadva, ezért nem történt utdíj kalkulálás.")
            return

        if not self.ugyfel.get("Város") or not self.ugyfel.get("Lakcím"):
            print("Nincs cím megadva, ezért nem történt utdíj kalkulálás.")
            return

        plusztav = 0
        for bolt in self.szabjegyszerk.boltok:
            plusztav = f"{plusztav}+{boltoktav[bolt]}"

        utdij = adatgyujto.Utdij_kalkulator(self.ugyfel.get("Város") + ", "+ self.ugyfel.get("Lakcím"))
        fogyasztas = utdij.utdij_kalkulacio(plusztav)

        sorok = [
            {
                "Anyag": "Benzin",
                "Szin": "",
                "URL": utdij.nav_url,
                "Mennyiseg": f"({fogyasztas.get("literek", 0)})*2*{toblettszorzo}",#fogyasztas.get("literek", 0) * 2 * toblettszorzo,
                "Mertekegyseg": "l",
                "Egysegar": fogyasztas.get("literar_huf", 0),
                "Osszar": f"({fogyasztas.get("uzemanyag_koltseg_huf", 0)})*2*{toblettszorzo}", #fogyasztas.get("uzemanyag_koltseg_huf", 0) * 2 * toblettszorzo,
            },
            {
                "Anyag": "Autóamortizáció",
                "Szin": "",
                "URL": "",
                "Mennyiseg": f"({fogyasztas.get("tavolsag_km", 0)})*2*{toblettszorzo}", #fogyasztas.get("tavolsag_km", 0) * 2 * toblettszorzo,
                "Mertekegyseg": "km",
                "Egysegar": fogyasztas.get("amortizacio_per_km", 0),
                "Osszar": f"({fogyasztas.get("auto_amortizacio_huf", 0)})*2*{toblettszorzo}",#fogyasztas.get("auto_amortizacio_huf", 0) * 2 * toblettszorzo,
            }
        ]

        anyagjegyzek_df = pd.DataFrame(sorok,
                                       columns=["Anyag", "Szin", "URL", "Mennyiseg", "Mertekegyseg", "Egysegar", "Osszar"])

        self.tablazat_tolto(anyagjegyzek_df)

    def reszosszegszamitas(self):
        if self.sor - 1 < self.kezdosor:
            return  # nincs mit összegezni

        cs = self.ws.cell(
            row=self.sor,
            column=9,
            value=f"=SUM(I{self.kezdosor}:I{self.sor - 1})"
        )
        cs.number_format = '#,##0 "Ft"'

    def anyagok_beirasa(self):
        anyagjegyzek, nyomtathato_szabjegy = self.szabjegyszerk.anyagjegyzek_szamitasa()
        print(self.szabjegyszerk.boltok)

        self.tablazat_tolto(anyagjegyzek)

        ws_2 = self.wb.create_sheet(title="szabasjegyzek")

        for i, col in enumerate(nyomtathato_szabjegy.columns, start=1):
            ws_2.cell(row=1, column=i, value=col)  # oszlopnevek

        for r in nyomtathato_szabjegy.itertuples(index=False):
            ws_2.append(r)

        self.utdij_beirasa()
        self.reszosszegszamitas()

    def _munkadij_mennyiseg_szamitas(self, munkanev: str, szamtip: str) -> float:
        """
        Munkadíj mennyiség számítása az anyagigény DataFrame-ből.
        Azonos Anyag különböző Színeit ÖSSZEADJA.
        """

        # Óra külön kezelés (később bővíthető)
        if szamtip == "Ora":
            return 0.0

        # érintett anyagok (ANYAG név szerint)
        erintett_anyagok = self.munkad_erintettanyag.get(munkanev, [])
        if not erintett_anyagok:
            return 0.0

        df = self.szabjegyszerk.anyagigeny
        if df is None or df.empty:
            return 0.0

        # ha nincs ilyen oszlop
        if szamtip not in df.columns:
            return 0.0

        # csak az érintett ANYAGOK (színtől függetlenül!)
        mask = df["Anyag"].astype(str).isin(erintett_anyagok)

        try:
            return float(
                df.loc[mask, szamtip]
                .fillna(0)
                .sum()
            )
        except Exception:
            return 0.0

    def munkadij_beiras(self):
        sorok = []

        mertek_dict = self.szabjegyszerk.mertekegyseg_dict

        for sor in self.munkadijlepesek:
            nev = sor.get("nev", "")
            leiras = sor.get("leiras", "")
            szamtip = sor.get("szamtip", "")
            ar = sor.get("ar", 0)

            # mértékegység
            mertekegyseg = mertek_dict.get(szamtip, "")

            # egységár
            try:
                ar = float(ar)
            except Exception:
                ar = 0.0

            # mennyiség számítás az anyagigényből
            menny = self._munkadij_mennyiseg_szamitas(nev, szamtip)

            # összár
            osszar = menny * ar

            sorok.append({
                "Anyag": nev,
                "Szin": leiras,
                "URL": "",
                "Mennyiseg": menny,
                "Mertekegyseg": mertekegyseg,
                "Egysegar": ar,
                "Osszar": osszar
            })

        if not sorok:
            print("ℹ️ Nincs munkadíj lépés.")
            return

        self.sor += 4
        self.kezdosor = self.sor

        munkajegyzek_df = pd.DataFrame(
            sorok,
            columns=[
                "Anyag", "Szin", "URL",
                "Mennyiseg", "Mertekegyseg",
                "Egysegar", "Osszar"
            ]
        )

        self.tablazat_tolto(munkajegyzek_df)
        self.reszosszegszamitas()

    def vegosszegszamitas(self):
        cs = self.ws.cell(
            row=self.sor+2,
            column=8,
            value=f"=I{self.kezdosor-4}+I{self.sor}"
        )
        cs.number_format = '#,##0 "Ft"'

        #egyesites
        self.ws.merge_cells(
            start_row=self.sor+2,
            start_column=8,
            end_row=self.sor+2,
            end_column=9
        )
        self.ws.merge_cells(
            start_row=self.sor + 2,
            start_column=6,
            end_row=self.sor + 2,
            end_column=7
        )
        self.ws.merge_cells(
            start_row=self.sor + 4,
            start_column=1,
            end_row=self.sor + 4,
            end_column=9
        )
        self.ws.merge_cells(
            start_row=self.sor + 7,
            start_column=1,
            end_row=self.sor + 7,
            end_column=9
        )
        self.ws.merge_cells(
            start_row=self.sor + 9,
            start_column=1,
            end_row=self.sor + 9,
            end_column=2
        )
        self.ws.merge_cells(
            start_row=self.sor + 10,
            start_column=1,
            end_row=self.sor + 10,
            end_column=9
        )

    def elkeszites(self):
        #self.kep_illeszto("fejlec.png","A1", [200, 100])
        if self.df is None:
            self.adatbeolvaso()
        self.szemelyes_adatok_onedrive()
        self.anyagok_beirasa()
        self.munkadij_beiras()
        self.vegosszegszamitas()

        self.faj.exel_kiiratasa(self.wb, self.faj.csv_utbol_arajanlat_ut())
